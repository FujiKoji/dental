import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
import openpyxl
import random

class Scraping:
    def __init__(self,id,pw,excel_path):
        self.id = id
        self.pw = pw
        self.excel_path = excel_path

    def get_url(self,page_num):
        driver_path = './chromedriver'
        options = Options()
        driver = webdriver.Chrome(executable_path=driver_path)
        driver.implicitly_wait(5)

        #ログインページに遷移
        url_login = "https://user.stransa.co.jp"
        driver.get(url_login)

        #ログイン
        xpath_id = '//*[@id="root"]/div/div/div/div/form/div[1]/div[2]/div[1]/div/div/div/input'
        driver.find_element(By.XPATH, xpath_id).send_keys(self.id)
        xpath_pw = '//*[@id="root"]/div/div/div/div/form/div[1]/div[2]/div[2]/div/div/input'
        driver.find_element(By.XPATH, xpath_pw).send_keys(self.pw)
        xpath_login = '//*[@id="root"]/div/div/div/div/form/div[2]/button'
        driver.find_element(By.XPATH, xpath_login).click()
        time.sleep(10)

        #url取得
        #患者一覧ページに遷移
        url_all_patient = 'https://apo-toolboxes.stransa.co.jp/user/patient/all'
        driver.get(url_all_patient)

        # #患者ごとのurlを取得する
        urls = []
        num = []

        for ii in range(page_num):
            for jj in range(50):
                #urlをリストに格納
                xpath_patient_list = f'//*[@id="patient-list-app"]/div/div[2]/ul/li[{jj+1}]/div[1]/div/p/span[1]/a'
                xpath_patient_num = f'//*[@id="patient-list-app"]/div/div[2]/ul/li[{jj+1}]/div[1]/div/ul/li[3]'
                urls.append(driver.find_element(By.XPATH, xpath_patient_list).get_attribute("href"))
                num.append(driver.find_element(By.XPATH, xpath_patient_num).text)

            #次のページに遷移
            xpath_next = '//*[@id="patient-list-app"]/section/div[2]/div[2]/ul/li[2]/a'
            driver.find_element(By.XPATH, xpath_next).click()
            time.sleep(3)
        driver.quit()
        return urls, num

    def get_treatmentdata(self,start,end):
        driver_path = './chromedriver'
        options = Options()
        driver = webdriver.Chrome(executable_path=driver_path)
        driver.implicitly_wait(5)

        #ログインページに遷移
        url_login = "https://user.stransa.co.jp"
        driver.get(url_login)

        #ログイン
        xpath_id = '//*[@id="root"]/div/div/div/div/form/div[1]/div[2]/div[1]/div/div/div/input'
        driver.find_element(By.XPATH, xpath_id).send_keys(self.id)
        xpath_pw = '//*[@id="root"]/div/div/div/div/form/div[1]/div[2]/div[2]/div/div/input'
        driver.find_element(By.XPATH, xpath_pw).send_keys(self.pw)
        xpath_login = '//*[@id="root"]/div/div/div/div/form/div[2]/button'
        driver.find_element(By.XPATH, xpath_login).click()
        time.sleep(10)

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['url一覧']
        ws_2 = wb['全患者診療データ']

        urls = []
        for ii in range(start,end):
            a = ws.max_row
            urls.append(ws.cell(row=ii,column=2).value)

        for ii in range(len(urls)):
            driver.get(urls[ii])
            time.sleep(2+5*random.random())
            xpath_num = '//*[@id="app"]/section[1]/div[2]/div/div[1]/ul/li[1]/span'
            xpath_all = '//*[@id="app"]/section[2]/div[2]/div/div[2]/ul'
            for kk in range(len(driver.find_elements(By.XPATH, xpath_all))):
                info = []
                info.append(driver.find_element(By.XPATH, xpath_num).text)
                for jj in range(2,10):
                    xpath = f'//*[@id="app"]/section[2]/div[2]/div/div[2]/ul[{kk+1}]/li[{jj}]'
                    info.append(driver.find_element(By.XPATH, xpath).text)
                a=ws_2.max_row
                for jj in range(len(info)):
                    ws_2.cell(row=a+1,column=jj+1).value = info[jj]
            time.sleep(1+3*random.random())
            print(ii)
        wb.save(self.excel_path)
        driver.quit()

class analysis:
    #スクレイピングしたデータに不備がないかチェック
    def check_scraping(self, inupt, output):
        wb = openpyxl.load_workbook(input)
        ws = wb['全患者詳細データ']
        used = ws['A':'A']
        used_data=[]
        for ii in range(len(used)):
            used_data.append(used[ii].value)
        ws = wb['url一覧']
        all = ws['A':'A']
        all_data=[]
        for ii in range(len(all)):
            all_data.append(all[ii].value)

        for ii in range(len(all_data)):
            if all_data[ii] in used_data:
                ws.cell(ii+1,3,value='ok')
        wb.save(output)

    #終了のみ患者データの予約内容を要素ごとに分解
    def preprocessing(self, inuput, output):
        wb = openpyxl.load_workbook(input)
        ws = wb['終了のみ患者診療データ']
        values = ws['B':'B']

        for ii, value in enumerate(values[1:]):
            data = []
            data.append(value.value.split('\n')[0])
            others = value.value.split('\n')[1]
            data.append(others.split('(')[0])
            data.append(others.split('(')[1].split(')')[0])
            data.append(others.split('(')[1].split(')')[1])
            data.append(others.split('(')[2].split(')')[0])
            
            for jj in range(3,8):
                data_range = ws.cell(ii+2,jj,value=data[jj-3])
        wb.save(output)
    