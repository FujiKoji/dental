from locale import DAY_1
from tokenize import Double3
import openpyxl

#スクレイピングしたデータに不備がないかチェック
def check_alldata_input():
    wb = openpyxl.load_workbook('./apotool_analysis_v4.xlsx')
    ws = wb['Sheet2']
    used = ws['A':'A']
    used_data=[]
    for ii in range(len(used)):
        used_data.append(used[ii].value)
    ws = wb['url一覧']
    all = ws['A':'A']
    all_data=[]
    for ii in range(len(all)):
        all_data.append(all[ii].value)
    
    for ii in range(len(all_data)):
        if all_data[ii] in used_data:
            ws.cell(ii+1,3,value='ok')
    
    wb.save('./apotool_analysis_v5.xlsx')

#来院回数を記載
def num_times():
    wb = openpyxl.load_workbook('./apotool_analysis_v3.xlsx')
    ws = wb['url一覧']
    urls_num = ws['A':'A']
    urls=[]
    for ii in range(len(urls_num)):
        urls.append(urls_num[ii].value)

    ws_1 = wb['全患者詳細データ']
    used_num = ws_1['A':'A']
    num=[]
    used_times = ws_1['V':'V']
    times=[]
    for ii in range(len(used_num)):
        num.append(used_num[ii].value)
        times.append(used_times[ii].value)
    
    for ii in range(len(urls)):
        if urls[ii] in num:
            row = num.index(urls[ii])
            ws.cell(ii+1,4,value=times[row])
    
    wb.save('./apotool_analysis_v4.xlsx')

#予約内容を要素ごとに分解
def preprocessing():
    wb = openpyxl.load_workbook('./apotool_analysis_v5.xlsx')
    ws = wb['Sheet1']
    values = ws['B':'B']

    for ii, value in enumerate(values[1:]):
        data = []
        data.append(value.value.split('\n')[0])
        others = value.value.split('\n')[1]
        data.append(others.split('(')[0])
        data.append(others.split('(')[1].split(')')[0])
        data.append(others.split('(')[1].split(')')[1])
        data.append(others.split('(')[2].split(')')[0])
        
        for jj in range(3,8):
            data_range = ws.cell(ii+2,jj,value=data[jj-3])
    wb.save('./apotool_analysis_v5.xlsx')

#来訪回数の導出
import datetime
def mark_treat_block():
    wb = openpyxl.load_workbook('./apotool_analysis_v11.xlsx')
    ws = wb['終了のみ患者データ']
    use_data_origin = wb['全患者一覧']['A':'A']
    use_data = []
    for data in use_data_origin:
        use_data.append(data.value)

    num = ws['A':'A'] #診察券番号
    date = ws['C':'C'] #診察日時
    flag = num[1].value
    date_before = date[1].value.split('/')
    date_before = datetime.date(int(date_before[0]),int(date_before[1]),int(date_before[2]))
    count = 1

    check = []

    for ii in range(1,len(num)):
        if num[ii].value == flag:
            if num[ii].value in use_data:
                date_now = date[ii].value.split('/')
                date_now = datetime.date(int(date_now[0]),int(date_now[1]),int(date_now[2]))
                days = (date_now - date_before).days
                date_before = date_now
                if days>=61:
                    count += 1
            else:
                continue
        elif num[ii].value != flag:
            if num[ii-1].value in use_data:
                wb['全患者一覧'].cell(use_data.index(flag)+1,3,value=count)

                check.append(num[ii-1].value)

                flag = num[ii].value
                date_before = date[ii].value.split('/')
                date_before = datetime.date(int(date_before[0]),int(date_before[1]),int(date_before[2]))
                print(count)
                count = 1
            else:
                flag = num[ii].value
    print(check)
    wb.save('./apotool_analysis_v12.xlsx')


#全患者詳細データから一部のデータを抽出する際に使用
def move_data():
    wb = openpyxl.load_workbook('./apotool_analysis_v10.xlsx')
    ws = wb['全患者詳細データ']
    ws_2 = wb['全患者一覧']

    nums = ws['A':'A']
    check = ws_2['A':'A']
    d1 = ws['J':'J']
    d2 = ws['K':'K']
    d3 = ws['V':'V']

    ch_list = []
    for ii in range(len(check)):
        ch_list.append(check[ii].value)

    for ii in range(1,len(nums)):
        if nums[ii].value in ch_list:
            index = ch_list.index(nums[ii].value)
            ws_2.cell(index+1,4,d1[ii].value)
            ws_2.cell(index+1,5,d2[ii].value)
            ws_2.cell(index+1,2,d3[ii].value)
    wb.save('./apotool_analysis_v11.xlsx')


#各月の初診人数計算
def count_get_patient():
    wb = openpyxl.load_workbook('./Book1.xlsx')
    ws = wb['Sheet2']
    ws_2 = wb['Sheet1']['B':'B']

    date = []
    for ii in range(2,30):
        date.append(ws.cell(1,ii).value)

    # 初診日
    get_date = []
    for ii in range(1,len(ws_2)):
        get_date.append(ws_2[ii].value)

    #最終来院日
    ws_2 = wb['Sheet1']['C':'C']
    out_date = []
    for ii in range(1,len(ws_2)):
        out_date.append(ws_2[ii].value)

    for ii in range(len(date)-1):
        count = 0
        for jj in range(len(get_date)):
            if  date[ii]<= get_date[jj] <date[ii+1]:
                count +=1
            
        for jj in range(ii+1,len(date)-1):
            out_count = 0
            for kk in range(len(get_date)):
                if  date[ii]<= get_date[kk] <date[ii+1]:
                    if out_date[kk]>=date[jj]:
                        out_count += 1
            ws.cell(ii+2,jj+2,value=out_count)

        ws.cell(ii+2,ii+2,value=count)

    wb.save('./Book2.xlsx')

wb = openpyxl.load_workbook('./Book2.xlsx')
ws = wb['Sheet2']
for ii in range(3,17):
    all = ws.cell(ii-1,ii).value
    next = ws.cell(ii-1,ii+12).value
    ws.cell(ii-1,1,value=next/all)
    
# for ii in range(3,4):
#     all = ws.cell(ii-1,ii-1).value
#     next = ws.cell(ii-1,ii+24).value
#     # ws.cell(ii,2,value=next/all)
#     print(next)
wb.save('./Book3.xlsx')

#データ範囲前から現在の患者数を推測
def count_patient_before():
    wb = openpyxl.load_workbook('./apotool_analysis_v13.xlsx')
    ws = wb['全患者一覧']["E":"E"]
    out = []
    for ii in range(1,len(ws)):
        out.append(ws[ii].value)
    ws = wb['患者数推移']

    date = []
    for ii in range(3,70):
        date.append(ws.cell(2,ii).value)

    for ii in range(len(date)):
        count = 0
        for jj in range(len(out)):
            if out[jj]>=date[ii]:
                count += 1
        ws.cell(3,3+ii,value=count)
    
    wb.save('./apotool_analysis_v14.xlsx')

import datetime
#各月の来院人数カウント
def count_visitors():
    wb = openpyxl.load_workbook('./apotool_analysis_v14.xlsx')
    ws = wb['終了のみ患者診療データ']["C":"C"]
    ws_w = wb['患者数推移_2']

    date_visit = []
    for ii in ws[1:]:
        a = ii.value.split("/")
        date_visit.append(datetime.datetime(int(a[0]),int(a[1]),int(a[2])))
    date_check = []
    for ii in range(2,68):
        date_check.append(ws_w.cell(1,ii).value)

    for ii in range(len(date_check)-1):
        count = 0
        for jj in range(1,len(date_visit)):
            if date_check[ii]<=date_visit[jj]<date_check[ii+1]:
                count += 1
        ws_w.cell(2,ii+2,value=count)
    wb.save('./apotool_analysis_v15.xlsx')

#アクティブ患者の来院頻度分析
def active_frequency():
    wb = openpyxl.load_workbook('./apotool_analysis_v20.xlsx')
    ws = wb['全患者一覧']
    ws_w = wb['患者数推移_2']

    visit_freq = []
    first_date = []
    last_date = []
    for ii in range(len(ws["A":"A"])):
        visit_freq.append(ws["G":"G"][ii].value)
        first_date.append(ws["D":"D"][ii].value)
        last_date.append(ws["E":"E"][ii].value)

    date = []
    for ii in range(2,68):
        date.append(ws_w.cell(1,ii).value)

    for ii in range(len(date)-1):
        count = [0]*13
        all_count=0
        ave = 0
        for jj in range(1,len(visit_freq)):
            if type(visit_freq[jj]) != str and type(visit_freq[jj]) != type(None) and type(first_date[jj])==type(datetime.datetime(2022,1,1)) and type(last_date[jj])==type(datetime.datetime(2022,1,1)):
                if first_date[jj] < date[ii+1]:
                    if last_date[jj] >= date[ii]:
                        for kk in range(13):
                            if ws_w.cell(6+kk,1).value <=visit_freq[jj] <=ws_w.cell(7+kk,1).value:
                                count[kk] += 1
                                all_count += 1
                                ave += visit_freq[jj]
        
        for jj in range(len(count)):
            ws_w.cell(6+jj,ii+2,value=count[jj])
            ws_w.cell(20,ii+2,value=ave/all_count)
            ws_w.cell(21,ii+2,value=all_count)
    wb.save('./apotool_analysis_v21.xlsx')

#来院・来訪頻度計算
def visit_frequency():
    wb = openpyxl.load_workbook('./apotool_analysis_v18.xlsx')
    ws = wb['全患者一覧']
    visit_num = []
    for ii in ws["B":"B"]:
        visit_num.append(ii.value)
    visit_block_num = []
    for ii in ws["C":"C"]:
        visit_block_num.append(ii.value)
    first_date = []
    for ii in ws["D":"D"]:
        first_date.append(ii.value)
    last_date = []
    for ii in ws["E":"E"]:
        last_date.append(ii.value)

    data_last_date = datetime.datetime(2017,4,17)

    for ii in range(1,len(first_date)):
        if type(visit_num[ii])!=type(None) and type(visit_block_num[ii])!=type(None) and type(first_date[ii])!=type(None) and type(last_date[ii])!=type(None) and first_date[ii]!=0 and last_date[ii]!=0:
            if first_date[ii]<=data_last_date:
                ws.cell(ii+1,7,value=(last_date[ii]-data_last_date).days/visit_num[ii])
                ws.cell(ii+1,8,value=(last_date[ii]-data_last_date).days/visit_block_num[ii])
            elif first_date[ii] > data_last_date:
                ws.cell(ii+1,7,value=(last_date[ii]-first_date[ii]).days/visit_num[ii])
                ws.cell(ii+1,8,value=(last_date[ii]-first_date[ii]).days/visit_block_num[ii])
    wb.save('./apotool_analysis_v19.xlsx')

#チェア稼働率導出
def work_time_per():
    wb = openpyxl.load_workbook('./apotool_analysis_v21.xlsx')
    ws = wb['終了のみ患者診療データ']
    ws_day = wb['チェア稼働率_日単位']
    date = []
    day_of_week = []
    time = []
    for ii in ws["C":"C"]:
        if ii.value!="日時":
            date.append(datetime.datetime(int(ii.value.split("/")[0]),int(ii.value.split("/")[1]),int(ii.value.split("/")[2])))
    for ii in ws["D":"D"]:
        if ii.value!="曜日":
            day_of_week.append(ii.value)
    for ii in ws["F":"F"]:
        if ii.value!="治療時間":
            time.append(int(ii.value))

    date_move = 0
    for ii in range(len(date)):
        if date_move == date[ii]:
            patient += 1
            work_time += time[ii]
        elif date_move != date[ii]:
            max_row = ws_day.max_row
            if max_row !=1:
                ws_day.cell(max_row,2,value=patient)
                ws_day.cell(max_row,4,value=work_time/60)
            date_move = date[ii]
            patient = 1
            work_time = time[ii]
            if day_of_week[ii]=="土":
                day_time = 6.5
            else:
                day_time = 8.5
            
            ws_day.cell(max_row+1,1,value=date_move)
            ws_day.cell(max_row+1,3,value=day_time)
    wb.save('./apotool_analysis_v22.xlsx')

#患者数推移_2に稼働率転記
def work_time_per_move():

wb = openpyxl.load_workbook('./apotool_analysis_v22.xlsx')
ws_day = wb['チェア稼働率_日単位']
date = []
for ii in ws_day["A":"A"]:
    date.append(ii.value)
per = []
for ii in ws_day["F":"F"]:
    per.append(ii.value)
ws = wb['患者数推移_2']
month = []
for ii in range(2,68):
    month.append(ws.cell(1,ii).value)

for ii in range(len(month)-1):
    count = 0
    per_count = 0
    for jj in range(1,len(date)):
        if month[ii]<=date[jj]<month[ii+1]:
            per_count += per[jj]
            count +=1
    value = per_count/count
    ws.cell(5,ii+2,value=value)
wb.save('./apotool_analysis_v23.xlsx')